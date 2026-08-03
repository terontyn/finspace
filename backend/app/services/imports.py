import codecs
import csv
import hashlib
import json
import re
import uuid
import zipfile
from collections.abc import Iterator
from datetime import UTC, date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.errors import ApiError
from app.db.models.accounts import Account
from app.db.models.categories import Category
from app.db.models.imports import ImportBatch, ImportRow
from app.db.models.transactions import FinancialTransaction
from app.dependencies.context import RequestContext
from app.schemas.imports import ImportMappingRequest, ImportRowOverrideRequest
from app.services.audit import record_audit, snapshot
from app.services.sync_outbox import enqueue_entity

MAPPING_FIELDS = {
    "date",
    "time",
    "transaction_type",
    "amount",
    "income_amount",
    "expense_amount",
    "currency",
    "account",
    "target_account",
    "category",
    "counterparty",
    "description",
    "comment",
    "status",
    "external_id",
}
TYPE_ALIASES = {
    "income": "income",
    "доход": "income",
    "приход": "income",
    "expense": "expense",
    "расход": "expense",
    "списание": "expense",
    "transfer": "transfer",
    "перевод": "transfer",
    "refund": "refund",
    "возврат": "refund",
    "adjustment": "adjustment",
    "корректировка": "adjustment",
}
STATUS_ALIASES = {
    "": "confirmed",
    "confirmed": "confirmed",
    "подтверждена": "confirmed",
    "draft": "draft",
    "черновик": "draft",
    "reconciled": "reconciled",
    "сверена": "reconciled",
}


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value).strip())


def _csv_encoding(path: Path) -> tuple[str, str]:
    for codec, label in (("utf-8-sig", "utf-8"), ("cp1251", "windows-1251")):
        decoder = codecs.getincrementaldecoder(codec)(errors="strict")
        try:
            with path.open("rb") as source:
                while chunk := source.read(64 * 1024):
                    if b"\x00" in chunk:
                        raise ApiError(
                            status_code=422,
                            code="IMPORT_FILE_TYPE_NOT_ALLOWED",
                            message="CSV contains binary data",
                        )
                    decoder.decode(chunk)
                decoder.decode(b"", final=True)
            return codec, label
        except UnicodeDecodeError:
            continue
    raise ApiError(
        status_code=422,
        code="IMPORT_FILE_TYPE_NOT_ALLOWED",
        message="CSV encoding is not supported",
    )


def _csv_rows(path: Path) -> tuple[Iterator[tuple[str | None, int, dict[str, str]]], str]:
    codec, encoding = _csv_encoding(path)
    with path.open("r", encoding=codec, newline="") as source:
        sample = source.read(8192)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t,")
    except csv.Error:
        dialect = None

    def iterator() -> Iterator[tuple[str | None, int, dict[str, str]]]:
        with path.open("r", encoding=codec, newline="") as source:
            reader = (
                csv.DictReader(source, dialect=dialect)
                if dialect is not None
                else csv.DictReader(source, dialect="excel", delimiter=";")
            )
            if not reader.fieldnames:
                raise ApiError(
                    status_code=422,
                    code="IMPORT_MAPPING_INVALID",
                    message="CSV has no header",
                )
            for row_number, row in enumerate(reader, start=2):
                yield (
                    None,
                    row_number,
                    {
                        _clean_text(key): _clean_text(value)
                        for key, value in row.items()
                        if key is not None
                    },
                )

    return iterator(), f"csv-{encoding}"


def _xlsx_rows(path: Path) -> tuple[Iterator[tuple[str | None, int, dict[str, str]]], str]:
    with path.open("rb") as source:
        if source.read(4) != b"PK\x03\x04":
            raise ApiError(
                status_code=422,
                code="IMPORT_FILE_TYPE_NOT_ALLOWED",
                message="XLSX signature is invalid",
            )
    try:
        with zipfile.ZipFile(path) as archive:
            unsafe_parts = {
                name
                for name in archive.namelist()
                if name.casefold().endswith("vbaproject.bin") or "/activex/" in name.casefold()
            }
            if unsafe_parts:
                raise ApiError(
                    status_code=415,
                    code="IMPORT_FILE_TYPE_NOT_ALLOWED",
                    message="Macro-enabled or ActiveX workbooks are not allowed",
                )
    except zipfile.BadZipFile as exc:
        raise ApiError(
            status_code=422,
            code="IMPORT_FILE_TYPE_NOT_ALLOWED",
            message="XLSX archive is invalid",
        ) from exc
    workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=False, keep_links=False)

    def iterator() -> Iterator[tuple[str | None, int, dict[str, str]]]:
        try:
            for sheet in workbook.worksheets:
                rows = sheet.iter_rows(values_only=True)
                headers = next(rows, None)
                if headers is None:
                    continue
                names = [_clean_text(value) for value in headers]
                for row_number, values in enumerate(rows, start=2):
                    yield (
                        sheet.title,
                        row_number,
                        {
                            name: _clean_text(value)
                            for name, value in zip(names, values, strict=False)
                            if name
                        },
                    )
        finally:
            workbook.close()

    return iterator(), "xlsx-journal"


async def upload_import(
    session: AsyncSession,
    context: RequestContext,
    upload: UploadFile,
    *,
    force_duplicate: bool,
) -> ImportBatch:
    original_name = Path(upload.filename or "").name
    extension = Path(original_name).suffix.lower().lstrip(".")
    if extension not in settings.allowed_import_extensions:
        raise ApiError(
            status_code=415,
            code="IMPORT_FILE_TYPE_NOT_ALLOWED",
            message="Only CSV and XLSX files are allowed",
        )
    storage = settings.import_storage_path.resolve()
    storage.mkdir(parents=True, exist_ok=True, mode=0o700)
    stored_filename = f"{uuid.uuid4().hex}.{extension}"
    stored_path = storage / stored_filename
    max_size = settings.import_max_file_size_mb * 1024 * 1024
    digest = hashlib.sha256()
    size = 0
    try:
        with stored_path.open("xb") as destination:
            while chunk := await upload.read(1024 * 1024):
                size += len(chunk)
                if size > max_size:
                    raise ApiError(
                        status_code=413,
                        code="IMPORT_FILE_TOO_LARGE",
                        message="Import file exceeds configured size limit",
                    )
                digest.update(chunk)
                destination.write(chunk)
        if size == 0:
            raise ApiError(
                status_code=422, code="IMPORT_FILE_TYPE_NOT_ALLOWED", message="File is empty"
            )
        file_hash = digest.hexdigest()
        duplicate = await session.scalar(
            select(ImportBatch.id).where(
                ImportBatch.workspace_id == context.workspace.id,
                ImportBatch.file_sha256 == file_hash,
                ImportBatch.status != "cancelled",
            )
        )
        if duplicate is not None and not force_duplicate:
            raise ApiError(
                status_code=409,
                code="IMPORT_DUPLICATE_FILE",
                message="This file was already uploaded",
                details={"batch_id": str(duplicate)},
            )
        rows, detected_format = (
            _csv_rows(stored_path) if extension == "csv" else _xlsx_rows(stored_path)
        )
        batch = ImportBatch(
            workspace_id=context.workspace.id,
            created_by=context.user.id,
            filename=original_name,
            stored_filename=stored_filename,
            file_type=extension,
            file_size=size,
            file_sha256=file_hash,
            status="uploaded",
            detected_format=detected_format,
        )
        session.add(batch)
        await session.flush()
        row_count = 0
        source_columns: set[str] = set()
        for source_sheet, row_number, raw_data in rows:
            row_count += 1
            if row_count > settings.import_max_rows:
                raise ApiError(
                    status_code=413,
                    code="IMPORT_ROW_LIMIT_EXCEEDED",
                    message="Import file exceeds configured row limit",
                )
            source_columns.update(raw_data)
            session.add(
                ImportRow(
                    batch_id=batch.id,
                    source_sheet=source_sheet,
                    source_row_number=row_number,
                    raw_data=raw_data,
                    status="raw",
                )
            )
            if row_count % 1000 == 0:
                await session.flush()
        batch.status = "mapping_required"
        batch.summary = {
            "total": row_count,
            "source_columns": sorted(source_columns),
            "valid": 0,
            "invalid": 0,
            "duplicate": 0,
            "skipped": 0,
        }
        await record_audit(
            session,
            workspace_id=context.workspace.id,
            actor_user_id=context.user.id,
            entity_type="import_batch",
            entity_id=batch.id,
            action="import.upload",
            before_data=None,
            after_data={"filename": original_name, "file_size": size, "rows": row_count},
            request_id=context.request_id,
        )
        await session.commit()
        await session.refresh(batch)
        return batch
    except Exception:
        await session.rollback()
        stored_path.unlink(missing_ok=True)
        raise


async def get_batch(
    session: AsyncSession, workspace_id: uuid.UUID, batch_id: uuid.UUID
) -> ImportBatch:
    batch = await session.scalar(
        select(ImportBatch).where(
            ImportBatch.id == batch_id,
            ImportBatch.workspace_id == workspace_id,
        )
    )
    if batch is None:
        raise ApiError(status_code=404, code="IMPORT_NOT_FOUND", message="Import was not found")
    return batch


async def override_duplicate(
    session: AsyncSession,
    context: RequestContext,
    batch_id: uuid.UUID,
    row_id: uuid.UUID,
    data: ImportRowOverrideRequest,
) -> ImportRow:
    batch = await get_batch(session, context.workspace.id, batch_id)
    if batch.status not in {"validated", "ready"}:
        raise ApiError(
            status_code=409,
            code="IMPORT_NOT_READY",
            message="The batch must be validated before overriding duplicates",
        )
    row = await session.scalar(
        select(ImportRow).where(ImportRow.id == row_id, ImportRow.batch_id == batch.id)
    )
    if row is None:
        raise ApiError(status_code=404, code="IMPORT_NOT_FOUND", message="Import row was not found")
    if row.status != "duplicate" or not data.import_as_new:
        raise ApiError(
            status_code=409,
            code="IMPORT_NOT_READY",
            message="Only a detected duplicate can be explicitly imported as new",
        )
    row.status = "valid"
    row.duplicate_transaction_id = None
    summary = dict(batch.summary or {})
    summary["duplicate"] = max(0, int(summary.get("duplicate", 0)) - 1)
    summary["valid"] = int(summary.get("valid", 0)) + 1
    batch.summary = summary
    batch.status = "ready"
    await record_audit(
        session,
        workspace_id=context.workspace.id,
        actor_user_id=context.user.id,
        entity_type="import_batch",
        entity_id=batch.id,
        action="import.validate",
        before_data=None,
        after_data={"row_id": str(row.id), "duplicate_override": True},
        request_id=context.request_id,
        source="import",
    )
    await session.commit()
    await session.refresh(row)
    return row


async def set_mapping(
    session: AsyncSession,
    context: RequestContext,
    batch_id: uuid.UUID,
    data: ImportMappingRequest,
) -> ImportBatch:
    batch = await get_batch(session, context.workspace.id, batch_id)
    unknown = set(data.mapping) - MAPPING_FIELDS
    required = {"date", "account"}
    has_amount = "amount" in data.mapping or {
        "income_amount",
        "expense_amount",
    }.intersection(data.mapping)
    if unknown or not required.issubset(data.mapping) or not has_amount:
        raise ApiError(
            status_code=422,
            code="IMPORT_MAPPING_INVALID",
            message="Import mapping is incomplete or contains unknown fields",
            details={"unknown": sorted(unknown), "required": sorted(required)},
        )
    batch.mapping = {"fields": data.mapping, "locale": data.locale}
    batch.status = "parsed"
    await record_audit(
        session,
        workspace_id=context.workspace.id,
        actor_user_id=context.user.id,
        entity_type="import_batch",
        entity_id=batch.id,
        action="import.mapping",
        before_data=None,
        after_data={"mapped_fields": sorted(data.mapping)},
        request_id=context.request_id,
    )
    await session.commit()
    await session.refresh(batch)
    return batch


def _mapped(raw: dict[str, Any], fields: dict[str, str], name: str) -> str:
    source = fields.get(name)
    return _clean_text(raw.get(source)) if source else ""


def _parse_decimal(value: str) -> Decimal:
    compact = value.replace("\u00a0", "").replace(" ", "")
    if "," in compact and "." in compact:
        if compact.rfind(",") > compact.rfind("."):
            compact = compact.replace(".", "").replace(",", ".")
        else:
            compact = compact.replace(",", "")
    else:
        compact = compact.replace(",", ".")
    try:
        amount = Decimal(compact)
    except InvalidOperation as exc:
        raise ValueError("Amount is not a decimal number") from exc
    if amount <= 0:
        raise ValueError("Amount must be positive")
    return amount.quantize(Decimal("0.0001"))


def _parse_occurred(date_value: str, time_value: str, timezone_name: str) -> datetime:
    parsed_date: date | None = None
    for date_format in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            parsed_date = datetime.strptime(date_value, date_format).date()
            break
        except ValueError:
            continue
    if parsed_date is None:
        try:
            parsed = datetime.fromisoformat(date_value.replace("Z", "+00:00"))
            if parsed.tzinfo is not None:
                return parsed.astimezone(UTC)
            parsed_date = parsed.date()
        except ValueError as exc:
            raise ValueError("Date format is not supported") from exc
    parsed_time = time()
    if time_value:
        for time_format in ("%H:%M:%S", "%H:%M"):
            try:
                parsed_time = datetime.strptime(time_value, time_format).time()
                break
            except ValueError:
                continue
        else:
            raise ValueError("Time format is not supported")
    try:
        timezone = ZoneInfo(timezone_name)
    except ZoneInfoNotFoundError as exc:
        raise ValueError("Workspace timezone is invalid") from exc
    return datetime.combine(parsed_date, parsed_time, tzinfo=timezone).astimezone(UTC)


def _name_map(items: list[Account] | list[Category]) -> dict[str, list[Any]]:
    result: dict[str, list[Any]] = {}
    for item in items:
        result.setdefault(item.name.strip().casefold(), []).append(item)
    return result


def _resolve_name(mapping: dict[str, list[Any]], value: str, entity: str) -> Any:
    matches = mapping.get(value.strip().casefold(), [])
    if not matches:
        raise ValueError(f"Unknown {entity}: {value}")
    if len(matches) > 1:
        raise ValueError(f"Ambiguous {entity}: {value}")
    return matches[0]


def _fingerprint(data: dict[str, Any]) -> str:
    fields = (
        data["workspace_id"],
        data["occurred_at"],
        data["transaction_type"],
        data["amount"],
        data["currency"],
        data["account_id"],
        data.get("target_account_id"),
        _clean_text(data.get("description")).casefold(),
        data.get("external_id"),
    )
    return hashlib.sha256(json.dumps(fields, ensure_ascii=False).encode("utf-8")).hexdigest()


async def _find_duplicate(
    session: AsyncSession, workspace_id: uuid.UUID, data: dict[str, Any]
) -> uuid.UUID | None:
    candidates = list(
        (
            await session.scalars(
                select(FinancialTransaction).where(
                    FinancialTransaction.workspace_id == workspace_id,
                    FinancialTransaction.deleted_at.is_(None),
                    FinancialTransaction.occurred_at == datetime.fromisoformat(data["occurred_at"]),
                    FinancialTransaction.transaction_type == data["transaction_type"],
                    FinancialTransaction.amount == Decimal(data["amount"]),
                    FinancialTransaction.currency == data["currency"],
                    FinancialTransaction.account_id == uuid.UUID(data["account_id"]),
                    FinancialTransaction.target_account_id
                    == (
                        uuid.UUID(data["target_account_id"])
                        if data.get("target_account_id")
                        else None
                    ),
                )
            )
        ).all()
    )
    normalized_description = _clean_text(data.get("description")).casefold()
    for candidate in candidates:
        if _clean_text(
            candidate.description
        ).casefold() == normalized_description and candidate.external_id == data.get("external_id"):
            return candidate.id
    return None


async def validate_import(
    session: AsyncSession, context: RequestContext, batch_id: uuid.UUID
) -> ImportBatch:
    batch = await get_batch(session, context.workspace.id, batch_id)
    if not batch.mapping or batch.status not in {"parsed", "validated", "ready"}:
        raise ApiError(
            status_code=409,
            code="IMPORT_MAPPING_INVALID",
            message="Mapping must be configured before validation",
        )
    mapping = batch.mapping
    fields = dict(mapping["fields"])
    accounts = list(
        (
            await session.scalars(
                select(Account).where(
                    Account.workspace_id == context.workspace.id,
                    Account.deleted_at.is_(None),
                    Account.is_archived.is_(False),
                )
            )
        ).all()
    )
    categories = list(
        (
            await session.scalars(
                select(Category).where(
                    Category.workspace_id == context.workspace.id,
                    Category.deleted_at.is_(None),
                    Category.is_archived.is_(False),
                )
            )
        ).all()
    )
    account_map = _name_map(accounts)
    category_map = _name_map(categories)
    rows = list(
        (
            await session.scalars(
                select(ImportRow)
                .where(ImportRow.batch_id == batch.id)
                .order_by(ImportRow.source_row_number)
            )
        ).all()
    )
    seen: set[str] = set()
    counts = {"total": len(rows), "valid": 0, "invalid": 0, "duplicate": 0, "skipped": 0}
    affected_accounts: set[str] = set()
    currencies: set[str] = set()
    dates: list[str] = []
    for row in rows:
        errors: list[dict[str, Any]] = []
        normalized: dict[str, Any] | None = None
        try:
            income_value = _mapped(row.raw_data, fields, "income_amount")
            expense_value = _mapped(row.raw_data, fields, "expense_amount")
            type_value = _mapped(row.raw_data, fields, "transaction_type").casefold()
            amount_value = _mapped(row.raw_data, fields, "amount")
            if income_value or expense_value:
                if income_value and expense_value:
                    raise ValueError("Income and expense columns cannot both contain a value")
                transaction_type = "income" if income_value else "expense"
                amount_value = income_value or expense_value
            else:
                transaction_type = TYPE_ALIASES.get(type_value, type_value)
            if transaction_type not in {"income", "expense", "transfer"}:
                raise ValueError("Transaction type is not supported for import")
            amount = _parse_decimal(amount_value)
            occurred_at = _parse_occurred(
                _mapped(row.raw_data, fields, "date"),
                _mapped(row.raw_data, fields, "time"),
                context.workspace.timezone,
            )
            account = _resolve_name(
                account_map, _mapped(row.raw_data, fields, "account"), "account"
            )
            target_account = None
            if transaction_type == "transfer":
                target_account = _resolve_name(
                    account_map,
                    _mapped(row.raw_data, fields, "target_account"),
                    "target account",
                )
                if target_account.id == account.id or target_account.currency != account.currency:
                    raise ValueError("Transfer accounts must be distinct and use the same currency")
            category = None
            category_value = _mapped(row.raw_data, fields, "category")
            if category_value:
                category = _resolve_name(category_map, category_value, "category")
                if transaction_type == "transfer":
                    raise ValueError("Transfer cannot contain a category")
            currency = (_mapped(row.raw_data, fields, "currency") or account.currency).upper()
            if currency != account.currency:
                raise ValueError("Currency does not match account currency")
            status_value = _mapped(row.raw_data, fields, "status").casefold()
            status = STATUS_ALIASES.get(status_value)
            if status is None:
                raise ValueError("Status is not supported")
            normalized = {
                "workspace_id": str(context.workspace.id),
                "occurred_at": occurred_at.isoformat(),
                "transaction_type": transaction_type,
                "amount": str(amount),
                "currency": currency,
                "account_id": str(account.id),
                "target_account_id": str(target_account.id) if target_account else None,
                "category_id": str(category.id) if category else None,
                "counterparty": _mapped(row.raw_data, fields, "counterparty") or None,
                "description": _mapped(row.raw_data, fields, "description") or None,
                "comment": _mapped(row.raw_data, fields, "comment") or None,
                "status": status,
                "external_id": _mapped(row.raw_data, fields, "external_id") or None,
            }
            fingerprint = _fingerprint(normalized)
            duplicate_id = await _find_duplicate(session, context.workspace.id, normalized)
            if fingerprint in seen or duplicate_id is not None:
                row.status = "duplicate"
                row.duplicate_transaction_id = duplicate_id
                counts["duplicate"] += 1
            else:
                row.status = "valid"
                counts["valid"] += 1
            seen.add(fingerprint)
            affected_accounts.add(account.name)
            if target_account:
                affected_accounts.add(target_account.name)
            currencies.add(currency)
            dates.append(occurred_at.date().isoformat())
        except ValueError as exc:
            errors.append({"code": "VALIDATION_ERROR", "message": str(exc)})
            row.status = "invalid"
            counts["invalid"] += 1
        row.normalized_data = normalized
        row.validation_errors = errors or None
    batch.status = "ready" if counts["valid"] > 0 else "validated"
    batch.summary = {
        **counts,
        "accounts": sorted(affected_accounts),
        "currencies": sorted(currencies),
        "date_from": min(dates) if dates else None,
        "date_to": max(dates) if dates else None,
    }
    await record_audit(
        session,
        workspace_id=context.workspace.id,
        actor_user_id=context.user.id,
        entity_type="import_batch",
        entity_id=batch.id,
        action="import.validate",
        before_data=None,
        after_data=counts,
        request_id=context.request_id,
    )
    await session.commit()
    await session.refresh(batch)
    return batch


async def commit_import(
    session: AsyncSession,
    context: RequestContext,
    batch_id: uuid.UUID,
    *,
    confirmation: bool,
    idempotency_key: str | None,
) -> tuple[ImportBatch, int]:
    batch = await get_batch(session, context.workspace.id, batch_id)
    if batch.status == "imported" and batch.idempotency_key == idempotency_key:
        imported_count = int(
            await session.scalar(
                select(func.count())
                .select_from(ImportRow)
                .where(ImportRow.batch_id == batch.id, ImportRow.status == "imported")
            )
            or 0
        )
        return batch, imported_count
    if batch.status == "imported":
        raise ApiError(
            status_code=409,
            code="IMPORT_ALREADY_COMMITTED",
            message="Import batch is already committed",
        )
    if not confirmation or not idempotency_key or batch.status != "ready":
        raise ApiError(
            status_code=409,
            code="IMPORT_NOT_READY",
            message="Validated import and explicit confirmation are required",
        )
    batch.status = "importing"
    batch.idempotency_key = idempotency_key
    rows = list(
        (
            await session.scalars(
                select(ImportRow).where(
                    ImportRow.batch_id == batch.id,
                    ImportRow.status == "valid",
                )
            )
        ).all()
    )
    for row in rows:
        data = row.normalized_data or {}
        transaction = FinancialTransaction(
            workspace_id=context.workspace.id,
            occurred_at=datetime.fromisoformat(str(data["occurred_at"])),
            transaction_type=str(data["transaction_type"]),
            amount=Decimal(str(data["amount"])),
            currency=str(data["currency"]),
            account_id=uuid.UUID(str(data["account_id"])),
            target_account_id=(
                uuid.UUID(str(data["target_account_id"])) if data.get("target_account_id") else None
            ),
            category_id=(uuid.UUID(str(data["category_id"])) if data.get("category_id") else None),
            counterparty=data.get("counterparty"),
            description=data.get("description"),
            comment=data.get("comment"),
            status=str(data["status"]),
            source="import",
            external_id=data.get("external_id"),
            created_by=context.user.id,
            updated_by=context.user.id,
            import_batch_id=batch.id,
        )
        session.add(transaction)
        await session.flush()
        row.created_transaction_id = transaction.id
        row.status = "imported"
        await record_audit(
            session,
            workspace_id=context.workspace.id,
            actor_user_id=context.user.id,
            entity_type="transaction",
            entity_id=transaction.id,
            action="create",
            before_data=None,
            after_data=snapshot("transaction", transaction),
            request_id=context.request_id,
            source="import",
        )
        await enqueue_entity(
            session,
            workspace_id=context.workspace.id,
            entity_type="transaction",
            entity=transaction,
        )
    batch.status = "imported"
    batch.confirmed_at = datetime.now(UTC)
    await record_audit(
        session,
        workspace_id=context.workspace.id,
        actor_user_id=context.user.id,
        entity_type="import_batch",
        entity_id=batch.id,
        action="import.commit",
        before_data=None,
        after_data={"created_transactions": len(rows)},
        request_id=context.request_id,
        source="import",
    )
    await session.commit()
    await session.refresh(batch)
    (settings.import_storage_path.resolve() / batch.stored_filename).unlink(missing_ok=True)
    return batch, len(rows)


async def rollback_import(
    session: AsyncSession,
    context: RequestContext,
    batch_id: uuid.UUID,
    *,
    force: bool,
) -> tuple[ImportBatch, int]:
    batch = await get_batch(session, context.workspace.id, batch_id)
    if batch.status == "rolled_back":
        return batch, 0
    if batch.status != "imported":
        raise ApiError(status_code=409, code="IMPORT_NOT_READY", message="Batch is not imported")
    transactions = list(
        (
            await session.scalars(
                select(FinancialTransaction).where(
                    FinancialTransaction.workspace_id == context.workspace.id,
                    FinancialTransaction.import_batch_id == batch.id,
                    FinancialTransaction.deleted_at.is_(None),
                )
            )
        ).all()
    )
    changed = [item for item in transactions if item.version != 1]
    if changed and not force:
        raise ApiError(
            status_code=409,
            code="IMPORT_ROLLBACK_CONFLICT",
            message="Imported transactions were changed after import",
            details={"transaction_ids": [str(item.id) for item in changed]},
        )
    now = datetime.now(UTC)
    for transaction in transactions:
        before = snapshot("transaction", transaction)
        transaction.deleted_at = now
        transaction.version += 1
        transaction.updated_by = context.user.id
        await record_audit(
            session,
            workspace_id=context.workspace.id,
            actor_user_id=context.user.id,
            entity_type="transaction",
            entity_id=transaction.id,
            action="delete",
            before_data=before,
            after_data=snapshot("transaction", transaction),
            request_id=context.request_id,
            source="import",
        )
        await enqueue_entity(
            session,
            workspace_id=context.workspace.id,
            entity_type="transaction",
            entity=transaction,
            operation="delete",
        )
    rows = list(
        (
            await session.scalars(
                select(ImportRow).where(
                    ImportRow.batch_id == batch.id,
                    ImportRow.status == "imported",
                )
            )
        ).all()
    )
    for row in rows:
        row.status = "rolled_back"
    batch.status = "rolled_back"
    batch.rolled_back_at = now
    await record_audit(
        session,
        workspace_id=context.workspace.id,
        actor_user_id=context.user.id,
        entity_type="import_batch",
        entity_id=batch.id,
        action="import.rollback",
        before_data=None,
        after_data={"rolled_back_transactions": len(transactions), "forced": force},
        request_id=context.request_id,
        source="import",
    )
    await session.commit()
    await session.refresh(batch)
    return batch, len(transactions)


async def cancel_import(
    session: AsyncSession, context: RequestContext, batch_id: uuid.UUID
) -> ImportBatch:
    batch = await get_batch(session, context.workspace.id, batch_id)
    if batch.status in {"imported", "rolled_back"}:
        raise ApiError(
            status_code=409, code="IMPORT_NOT_READY", message="Batch cannot be cancelled"
        )
    batch.status = "cancelled"
    await session.commit()
    await session.refresh(batch)
    (settings.import_storage_path.resolve() / batch.stored_filename).unlink(missing_ok=True)
    return batch
